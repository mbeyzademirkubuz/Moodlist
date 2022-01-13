from selenium import webdriver #selenium kütüphanesinden, linki açmak için chrome kullanılacağından, webdriver alt kütüphanesi import edilir
from selenium.webdriver.chrome.service import Service
import openpyxl #excel işlemleri için import edilir
from random import randint #random tam sayı seçimi için import edilir
import tkinter as tk #uygulama arayüzü (GUI) için tkinter kütüphanesi bundan sonra "tk" olarak tanımlanacak şekilde import edilir

s= Service("chromedriver.exe")
s1= Service("chromedriver2.exe")
s2= Service("chromedriver3.exe")

wb = openpyxl.load_workbook("sorular.xlsx")  # excel dosyasını import eder
ws = wb.active  # excel dosyasındaki aktiv(o an açık) sayfayı seçer
listeler = ["https://www.youtube.com/playlist?list=PLvgRGDFe9gpZ0smgCvH9SsRpewHC_WPL4", # müzik listelerinin linkleri tanımlanır
            "https://www.youtube.com/playlist?list=PLvgRGDFe9gpZ--tb2c-tVjHJdSYBDdlyn",
            "https://www.youtube.com/playlist?list=PLvgRGDFe9gpaxlosRsPNaVc4joyUqkRa6",
            "https://www.youtube.com/playlist?list=PLvgRGDFe9gpZ5l_7olU5V4qMAmR9-TNuv",
            "https://www.youtube.com/playlist?list=PLvgRGDFe9gpYJObuI_UcR2KjSnDrmKZEQ",
            "https://www.youtube.com/playlist?list=PLvgRGDFe9gpZlbWYcpVguz9QZW7P7k8hI",
            "https://www.youtube.com/playlist?list=PLvgRGDFe9gpa23lI-6v54nAUA7pSUxOiZ",
            "https://www.youtube.com/playlist?list=PLvgRGDFe9gpY4qVko9Ya91RZrz3gQ5nNo",
            "https://www.youtube.com/playlist?list=PLvgRGDFe9gpZdxdldnOGvvFf2LSgeg5Vc",
            "https://www.youtube.com/playlist?list=PLvgRGDFe9gpYm2NkPCtST4xeUIiG-NmR3"]

i = 0 #i için baştan tanımlanmış değer
sonuc = 0  #sonuclar için baştan tanımlanmış puan değeri
pencere = tk.Tk() #uygulama için pencere açılır
pencere.geometry('600x550') #uygulamanın pencere boyutları tanımlanır
pencere.title("Müzik Seçim Botu") #uygulamanın başlığı belirlenir

def tekrar(): #ana kod, programın bir butonla tekrar çalıştırılabilmesi için bir fonksyion olarak tanımlanır
    j = 1 #j için baştan tanımlanmış değer
    i = 0 #i için fonksiyonun içinde tekrardan tanımlanmış değer
    sonuc = 0 #sonuc için fonksiyonun içinde tekrardan tanımlanmış değer
    sayilar = [] #boş bir sayilar listesi açılır
    while j<11: #10 random sayı seçilmesi ve hiçbir random değerin aynı olmaması için döngü başlatılır
        intr = randint(1, 100) #1 ile 100 arasında random değer seçer
        if not intr in sayilar: #değerin daha önce seçilip seçilmediğini denetler
            sayilar.append(intr) #seçilmediyse listeye ekler
            j = j+1 #ve döngünün devamı için j değerini 1 artırır
    def a(): #butonlara atanan a komutu için fonksiyon tanımlanır
        nonlocal i, sonuc #i ve sonuc değerlerinin bu fonksyiona özel olmadığını belirtmek ve a fonksyionunda kullanmak için nonlocal ile tanımlanır
        def link(): #linke tıklamanın komut fonksiyonu tanımlanır
            for l in range(0, 10):  # 10 alternatif sonuc için döngü başlatır
                if sonuc > l * 8: #0 puan durumu harici için
                    link = listeler[l]  # sonuca göre listenin ilgili linki çekilir
                elif sonuc == 0:
                    link = listeler[0]  # 0 puan durumunda sonuc 0'dan büyük olamayacağı için özel şart belirlenir
            try:
                browser = webdriver.Chrome(service=s) #Chrome açılması için chromedriver'ı tanımlanır
            except:
                try: browser = webdriver.Chrome(service=s1) #eş zamanlı 3 chrome sürümü olabilir, 3 sürüm için ayrı ayrı chromedriver'lar tanımlanır
                except: browser = webdriver.Chrome(service= s2)
            browser.get(link)  # Listenin linki açılır
        if var.get() == "A": #butonun değeri kontrol edilir
            sonuc = sonuc + 8 #a şıkkının sonuca etkisi 8 kabul edilir
        elif var.get() == "B":
            sonuc = sonuc + 6 #şıkların değerleri sonuca eklenir
        elif var.get() == "C":
            sonuc = sonuc + 4
        elif var.get() == "D":
            sonuc = sonuc + 2
        elif var.get() == "E":
            sonuc = sonuc

        if i < 10: #10 soru için if'li döngü başlatılır
            sayi = sayilar[i] #random sayılardan ilki seçilir
            soru = ws['A' + str(sayi)].value  # excelin ilk sütunundan soruyu çeker
            asikki = ws['B' + str(sayi)].value  # excelin diğer sütunlarından cevapları çeker
            bsikki = ws['C' + str(sayi)].value
            csikki = ws['D' + str(sayi)].value
            dsikki = ws['E' + str(sayi)].value
            esikki = ws['F' + str(sayi)].value
            label.config(text=("Soru " + str(i+1) + " - " + soru)) #soruyu uygulamaya yansıtır
            butonA.config(text=asikki, value= "A") #şıkları değerleriyle birlikte uygulamaya yansıtır
            butonB.config(text=bsikki, value= "B")
            butonC.config(text=csikki, value= "C")
            butonD.config(text=dsikki, value= "D")
            butonE.config(text=esikki, value= "E")
            i = i + 1 #döngünün devamı için i değerine 1 ekler

        else: #10 soru sorulduktan sonra
            butonC.destroy() #buton C, D ve E ekrandan silinir
            butonD.destroy()
            butonE.destroy()
            label.config(text=("\nSonucunuz: "+ str(sonuc) + "/80")) #sonuc uygulamaya yansıtılır
            butonA.config(text="Linke Git!", command=link) #buton A, linke gitmek için link komutuyla güncellenir
            butonB.config(text="Testi Tekrar Yap!", command=tekrar) #buton B, tekrar denemek için tekrar komutuyla güncellenir

    label = tk.Label(pencere, text="Başlıyor...", height=5, width=66,font=("Ariel",12)) #uygulamanın başına, yüksekliği, genişliği, fontu ve puntosu belirtilen şekilde başlık yazar
    label.grid(row=0) #başlık uygun yere yerleştirişir
    var = tk.StringVar() #butonlarda kullanılacak variable değeri önceden tanımlanır
    butonA = tk.Radiobutton(pencere, text="BAŞLA", command=a, height=5, width=20, variable=var, bg="gray") #A şıkkı için arka plan rengi gri ve komutu a olan buton tanımlanır
    butonA.grid(row=2) #buton uygun yere yerleştirilir
    butonB = tk.Radiobutton(pencere, text="BAŞLA", command=a, height=5, width=20, variable=var, bg="gray") #diğer butonlara da aynı işlem uygulanır
    butonB.grid(row=3)
    butonC = tk.Radiobutton(pencere, text="BAŞLA", command=a, height=5, width=20, variable=var, bg="gray")
    butonC.grid(row=4)
    butonD = tk.Radiobutton(pencere, text="BAŞLA", command=a, height=5, width=20, variable=var, bg="gray")
    butonD.grid(row=5)
    butonE = tk.Radiobutton(pencere, text="BAŞLA", command=a, height=5, width=20, variable=var, bg="gray")
    butonE.grid(row=6)
    pencere.mainloop() #uygulamanın devamlı ekranda kalması için

tekrar() #ana kod için fonksiyon çağrılır